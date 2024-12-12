import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Função para carregar os dados
def carregar_dados(caminho_arquivo):
    try:
        dados = pd.read_excel(caminho_arquivo)
        return dados
    except Exception as e:
        raise Exception(f"Erro ao carregar o arquivo: {e}")

# Função para calcular frequência de cada número
def calcular_frequencia(dados):
    numeros = dados.iloc[:, 1:].values.flatten()
    frequencia = pd.Series(numeros).value_counts().sort_index()
    frequencia_df = pd.DataFrame({"Número": frequencia.index, "Frequência": frequencia.values})
    frequencia_df["% Frequência"] = (frequencia_df["Frequência"] / frequencia_df["Frequência"].sum()) * 100
    return frequencia_df

# Função para calcular pares e ímpares
def calcular_pares_impares(dados):
    numeros = dados.iloc[:, 1:]
    pares = numeros.applymap(lambda x: x % 2 == 0).sum().sum()
    impares = numeros.size - pares
    total = numeros.size
    return {"Pares": pares, "Ímpares": impares, "% Pares": pares / total * 100, "% Ímpares": impares / total * 100}

# Função para calcular a distribuição por décadas
def calcular_decadas(dados):
    numeros = dados.iloc[:, 1:].values.flatten()
    decadas = {f"{i}-{i+9}": 0 for i in range(1, 26, 10)}
    for numero in numeros:
        for decada in decadas.keys():
            inicio, fim = map(int, decada.split("-"))
            if inicio <= numero <= fim:
                decadas[decada] += 1
    total = sum(decadas.values())
    return {chave: (valor, valor / total * 100) for chave, valor in decadas.items()}

# Função para calcular repetições entre concursos
def calcular_repeticoes(dados):
    repeticoes = []
    for i in range(1, len(dados)):
        repetidos = len(set(dados.iloc[i - 1, 1:]).intersection(set(dados.iloc[i, 1:])))
        repeticoes.append(repetidos)
    repeticoes_df = pd.DataFrame({"Concurso": dados.iloc[1:, 0].values, "Repetidos": repeticoes})
    return repeticoes_df

# Função para selecionar os 25 números mais frequentes
def selecionar_numeros(frequencia_df):
    selecionados = frequencia_df.sort_values("Frequência", ascending=False).head(25)["Número"].values
    return selecionados

# Função para salvar os dados em um Excel
def salvar_em_excel(caminho_saida, frequencia_df, pares_impares, decadas, repeticoes_df, numeros_selecionados):
    wb = Workbook()

    # Aba Frequência
    ws_frequencia = wb.active
    ws_frequencia.title = "Frequência"
    for r in dataframe_to_rows(frequencia_df, index=False, header=True):
        ws_frequencia.append(r)

    # Aba Pares e Ímpares
    ws_pares_impares = wb.create_sheet(title="Pares e Ímpares")
    ws_pares_impares.append(["Categoria", "Quantidade", "%"])
    ws_pares_impares.append(["Pares", pares_impares["Pares"], pares_impares["% Pares"]])
    ws_pares_impares.append(["Ímpares", pares_impares["Ímpares"], pares_impares["% Ímpares"]])

    # Aba Décadas
    ws_decadas = wb.create_sheet(title="Décadas")
    ws_decadas.append(["Década", "Quantidade", "%"])
    for decada, (quantidade, porcentagem) in decadas.items():
        ws_decadas.append([decada, quantidade, porcentagem])

    # Aba Repetições
    ws_repeticoes = wb.create_sheet(title="Repetições")
    for r in dataframe_to_rows(repeticoes_df, index=False, header=True):
        ws_repeticoes.append(r)

    # Aba Números Selecionados
    ws_numeros = wb.create_sheet(title="Números Selecionados")
    ws_numeros.append(["Números Selecionados"])
    ws_numeros.append(list(numeros_selecionados))

    wb.save(caminho_saida)

# Função principal
def main():
    caminho_arquivo = r"D:\GitHub\Lottery\Lotofácil\resultados.xlsx"
    caminho_saida = r"D:\GitHub\Lottery\Lotofácil\analise_lotofacil.xlsx"

    try:
        # Carregar dados
        dados = carregar_dados(caminho_arquivo)

        # Análises
        frequencia_df = calcular_frequencia(dados)
        pares_impares = calcular_pares_impares(dados)
        decadas = calcular_decadas(dados)
        repeticoes_df = calcular_repeticoes(dados)
        numeros_selecionados = selecionar_numeros(frequencia_df)

        # Salvar resultados
        salvar_em_excel(caminho_saida, frequencia_df, pares_impares, decadas, repeticoes_df, numeros_selecionados)

        print(f"Análises concluídas! Resultados salvos em: {caminho_saida}")
    except Exception as e:
        print(f"Erro durante a execução: {e}")

if __name__ == "__main__":
    main()